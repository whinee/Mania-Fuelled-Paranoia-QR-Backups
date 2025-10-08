from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Parse JIS0208.TXT
mapping = {}

with open("data/JIS0208.TXT", encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        codepoints, comment = line.split("#")
        parts = codepoints.split()
        if len(parts) < 3:
            continue

        sjis_hex, jis_hex, uni_hex = parts[:3]
        jis = int(jis_hex, 16)
        unicode_cp = int(uni_hex, 16)
        char = chr(unicode_cp)

        # Convert to kuten form (row/column)
        kuten = jis - 0x2020
        ku = kuten >> 8
        ten = kuten & 0xFF
        mapping[(ku, ten)] = char

# Create Excel workbook
wb = Workbook()
ws = wb.create_sheet(title = "JIS X 0208 Table")

# Headers
ws["A1"] = "KU/TEN"
for ten in range(1, 95):
    ws.cell(row=1, column=ten + 1, value=ten)

for ku in range(1, 95):
    ws.cell(row=ku + 1, column=1, value=ku)
    for ten in range(1, 95):
        char = mapping.get((ku, ten), "")
        ws.cell(row=ku + 1, column=ten + 1, value=char)

# Optional: make cells square-ish
for col in range(2, 96):
    ws.column_dimensions[get_column_letter(col)].width = 3
for row in range(1, 96):
    ws.row_dimensions[row].height = 18

wb.save("data/jis_x_0208_table.xlsx")
print("Saved to data/jis_x_0208_table.xlsx ✅")
