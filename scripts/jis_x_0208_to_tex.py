# jis_excel_to_latex.py
#
# Reads a 94x94 Excel sheet of JIS X 0208 characters
# and outputs LaTeX tables split into 15x42 chunks.
#
# Usage: python3 jis_excel_to_latex.py jis.xlsx > jis_tables.tex

from openpyxl import load_workbook

# Config
ROW_BLOCK, COL_BLOCK = 23, 10 # split sizes
START_CODE = 0x21
END_CODE = 0x7E

OVERRIDE = {
    0x2121: "\\wiki[SP]{Space_character}",
}

def to_hex(idx):
    """Return two-digit hex string for JIS indices (21–7E)."""
    return f"{START_CODE + idx:02X}"

def make_table(sheet, row_start, row_end, col_start, col_end):
    if not any(
        sheet.cell(row=r + 2, column=c + 2).value
        for r in range(row_start, row_end)
        for c in range(col_start, col_end)
    ):
        return ""

    col_def = "|Sc" * (col_end - col_start + 2) + "|"
    out = []

    out.append("\\begin{table}[H]")
    out.append("\\Fontified")
    out.append("\\centering")
    out.append(
        f"\\caption{{Shift JIS X 0208: {to_hex(row_start)}–{to_hex(row_end - 1)} × {to_hex(col_start)}–{to_hex(col_end - 1)}}}"
    )
    out.append("\\scalebox{1.2}{")
    out.append(f"\\begin{{tabular}}{{{col_def}}}")
    out.append("\\hline")

    # header row
    header = " & " + " & ".join(
        [f"\\textbf{{{to_hex(c)}}}" for c in range(col_start, col_end)]
    ) + " \\\\ \\hline"
    out.append(header)

    # body
    for r in range(row_start, row_end):
        row_cells = [f"\\textbf{{{to_hex(r)}}}"]
        for c in range(col_start, col_end):
            code = ((START_CODE + r) << 8) | (START_CODE + c)
            val = OVERRIDE.get(code) or sheet.cell(row=r + 2, column=c + 2).value or ""
            row_cells.append(str(val))
        out.append(" & ".join(row_cells) + " \\\\ \\hline")

    out.append("\\end{tabular}")
    out.append("}")
    out.append("\\end{table}")
    out.append("")
    return "\n".join(out)


def main():
    # if len(sys.argv) < 2:
    #     print("Usage: python3 jis_excel_to_latex.py jis.xlsx")
    #     return

    wb = load_workbook("data/jis_x_0208_table.xlsx")
    sheet = wb.active  # assume grid is in the first sheet

    rows, cols = END_CODE - START_CODE + 1, END_CODE - START_CODE + 1

    for row_start in range(0, rows, ROW_BLOCK):
        for col_start in range(0, cols, COL_BLOCK):
            row_end = min(row_start + ROW_BLOCK, rows)
            col_end = min(col_start + COL_BLOCK, cols)
            print(make_table(sheet, row_start, row_end, col_start, col_end))

if __name__ == "__main__":
    main()
